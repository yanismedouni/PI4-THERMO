#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Calcul de la MAE et RMSE à partir des CSV de désagrégation.

Lit tous les fichiers resultats_desagregation_*.csv dans output/
et produit un tableau récapitulatif + un fichier Excel.

Usage :
    python src/calcul_metriques.py
    python src/calcul_metriques.py --dossier output/mes_resultats
    python src/calcul_metriques.py --rapport output/metriques.xlsx
"""

import argparse
import glob
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────
PROJECT_DIR = Path(__file__).resolve().parent.parent
OUTPUT_DIR  = PROJECT_DIR / "output"

COL_REEL    = "P_reel_clim"
COL_ESTIME  = "P_estime_clim"
COL_ON_PRED = "o_climatisation"
COL_ON_REEL = "P_reel_clim"     # binarisé avec seuil


# ─────────────────────────────────────────────────────────────────────────────
# CALCUL DES MÉTRIQUES
# ─────────────────────────────────────────────────────────────────────────────

def calculer_metriques(csv_path: str, seuil_on: float = 0.05) -> dict:
    """
    Calcule MAE, RMSE, Energy Fraction et Norm Error
    à partir d'un CSV de désagrégation.
    """
    df = pd.read_csv(csv_path)

    colonnes_requises = [COL_REEL, COL_ESTIME, COL_ON_PRED]
    manquantes = [c for c in colonnes_requises if c not in df.columns]
    if manquantes:
        raise ValueError(f"Colonnes manquantes dans {Path(csv_path).name} : {manquantes}")

    y_reel   = pd.to_numeric(df[COL_REEL],   errors="coerce").fillna(0).to_numpy()
    y_estime = pd.to_numeric(df[COL_ESTIME], errors="coerce").fillna(0).to_numpy()
    o_pred   = pd.to_numeric(df[COL_ON_PRED],errors="coerce").fillna(0).to_numpy()
    o_reel   = (y_reel > seuil_on).astype(int)

    n = len(df)

    # ── Métriques de puissance ────────────────────────────────────────
    rmse = float(np.sqrt(np.mean((y_reel - y_estime) ** 2)))
    mae  = float(np.mean(np.abs(y_reel - y_estime)))

    # ── Fraction d'énergie ────────────────────────────────────────────
    energie_reelle  = y_reel.sum()
    energie_estimee = y_estime.sum()
    energy_frac = float(energie_estimee / max(energie_reelle, 1e-9))

    # ── Erreur normalisée L1 ──────────────────────────────────────────
    norm_err = float(np.abs(y_reel - y_estime).sum() / max(np.abs(y_reel).sum(), 1e-9))

    # ── Métriques ON/OFF ─────────────────────────────────────────────
    o_pred_bin = (np.round(o_pred)).astype(int)
    tp = int(((o_reel == 1) & (o_pred_bin == 1)).sum())
    tn = int(((o_reel == 0) & (o_pred_bin == 0)).sum())
    fp = int(((o_reel == 0) & (o_pred_bin == 1)).sum())
    fn = int(((o_reel == 1) & (o_pred_bin == 0)).sum())

    recall    = tp / (tp + fn) if (tp + fn) > 0 else float("nan")
    precision = tp / (tp + fp) if (tp + fp) > 0 else float("nan")
    f1 = (2 * precision * recall / (precision + recall)
          if not (np.isnan(precision) or np.isnan(recall))
          and (precision + recall) > 0
          else float("nan"))
    fpr = fp / (fp + tn) if (fp + tn) > 0 else float("nan")

    # ── Gap relatif MOSEK ─────────────────────────────────────────────
    gap_reel = float("nan")
    if "gap_reel" in df.columns:
        vals = pd.to_numeric(df["gap_reel"], errors="coerce").dropna()
        if not vals.empty:
            gap_reel = float(vals.iloc[0])

    # ── Info sur le fichier ───────────────────────────────────────────
    # Extraire dataid et date depuis le nom du fichier
    # Format : resultats_desagregation_<dataid>_<date>_7jours.csv
    nom = Path(csv_path).stem   # ex. resultats_desagregation_3864_2015-07-02_7jours
    parties = nom.split("_")
    try:
        dataid     = parties[2]
        date_debut = parties[3]
    except IndexError:
        dataid     = "—"
        date_debut = "—"

    return {
        "Fichier"       : Path(csv_path).name,
        "Client (dataid)": dataid,
        "Date début"    : date_debut,
        "N pas"         : n,
        "RMSE (kW)"     : round(rmse, 4),
        "MAE (kW)"      : round(mae,  4),
        "Energy Frac"   : round(energy_frac, 4),
        "Norm Err"      : round(norm_err, 4),
        "Recall"        : round(recall,    4) if not np.isnan(recall)    else float("nan"),
        "Precision"     : round(precision, 4) if not np.isnan(precision) else float("nan"),
        "F1"            : round(f1,        4) if not np.isnan(f1)        else float("nan"),
        "FPR"           : round(fpr,       4) if not np.isnan(fpr)       else float("nan"),
        "TP"            : tp,
        "TN"            : tn,
        "FP"            : fp,
        "FN"            : fn,
        "Gap réel (%)" : round(gap_reel * 100, 2) if not np.isnan(gap_reel) else float("nan"),
    }


# ─────────────────────────────────────────────────────────────────────────────
# RAPPORT EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def sauvegarder_excel(resultats: list[dict], out_path: str) -> None:
    """Génère un rapport Excel formaté à partir des métriques calculées."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Métriques"
    ws.sheet_view.showGridLines = False

    # ── Styles ────────────────────────────────────────────────────────
    BLUE_DARK  = "1F3864"
    BLUE_MID   = "2E75B6"
    BLUE_LIGHT = "D6E4F0"
    WHITE      = "FFFFFF"
    GREY       = "F2F2F2"

    def fill(c): return PatternFill("solid", start_color=c)
    def font_(bold=False, color="000000", size=10):
        return Font(name="Arial", bold=bold, color=color, size=size)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin   = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Titre ─────────────────────────────────────────────────────────
    n_cols = 17
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    ws["A1"] = "THERMO — Métriques de désagrégation (Climatisation)"
    ws["A1"].font      = font_(bold=True, color=WHITE, size=13)
    ws["A1"].fill      = fill(BLUE_DARK)
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    # ── En-têtes ──────────────────────────────────────────────────────
    headers = [
        "Fichier CSV", "Client", "Date début", "N pas",
        "RMSE (kW)", "MAE (kW)", "Energy Frac", "Norm Err",
        "Recall", "Precision", "F1", "FPR",
        "Gap réel (%)",
        "TP", "TN", "FP", "FN",
    ]
    ws.row_dimensions[2].height = 32
    for col_idx, h in enumerate(headers, start=1):
        cell           = ws.cell(row=2, column=col_idx, value=h)
        cell.font      = font_(bold=True, color=WHITE, size=10)
        cell.fill      = fill(BLUE_MID)
        cell.alignment = center
        cell.border    = border

    # ── Lignes de données ─────────────────────────────────────────────
    keys = [
        "Fichier", "Client (dataid)", "Date début", "N pas",
        "RMSE (kW)", "MAE (kW)", "Energy Frac", "Norm Err",
        "Recall", "Precision", "F1", "FPR",
        "Gap réel (%)",
        "TP", "TN", "FP", "FN",
    ]
    for i, row_data in enumerate(resultats, start=3):
        shade = GREY if i % 2 == 0 else WHITE
        ws.row_dimensions[i].height = 18

        for col_idx, key in enumerate(keys, start=1):
            val = _cel(row_data.get(key, "—"))
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.fill      = fill(shade)
            cell.border    = border
            cell.font      = font_(size=10)
            cell.alignment = left if col_idx == 1 else center

    # ── Ligne de résumé (moyennes) ────────────────────────────────────
    if resultats:
        row_moy = len(resultats) + 3
        ws.row_dimensions[row_moy].height = 20

        ws.cell(row=row_moy, column=1, value="MOYENNE").font = font_(bold=True, color=WHITE, size=10)
        ws.cell(row=row_moy, column=1).fill      = fill(BLUE_DARK)
        ws.cell(row=row_moy, column=1).alignment = center
        ws.cell(row=row_moy, column=1).border    = border

        for col_idx, key in enumerate(keys[1:], start=2):
            vals = [float(r[key]) for r in resultats
                    if isinstance(r.get(key), (int, float, np.floating))
                    and not (r[key] != r[key])]   # NaN check universel
            val  = round(float(np.mean(vals)), 4) if vals else "—"
            cell = ws.cell(row=row_moy, column=col_idx, value=_cel(val))
            cell.fill      = fill(BLUE_LIGHT)
            cell.border    = border
            cell.font      = font_(bold=True, size=10)
            cell.alignment = center

    # ── Largeurs colonnes ─────────────────────────────────────────────
    col_widths = [42, 10, 12, 8, 10, 10, 12, 10, 10, 11, 10, 10, 12, 7, 7, 7, 7]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"

    wb.save(out_path)


# ─────────────────────────────────────────────────────────────────────────────
# ÉVALUATION SUR UNE PÉRIODE COMPLÈTE (résultats_ID_annee_####_complet.csv)
# ─────────────────────────────────────────────────────────────────────────────

def evaluer_periode(
    csv_path: str,
    seuil_on: float = 0.05,
    par_saison: bool = True,
    par_mois: bool = True,
) -> dict:
    """
    Évalue les performances sur un fichier consolidé annuel/été
    (format : resultats_<dataid>_<periode>_<annee>_complet.csv).

    Retourne un dict avec :
      - metriques_globales  : mêmes métriques que calculer_metriques()
      - metriques_par_saison: dict {saison: métriques} si par_saison=True
      - metriques_par_mois  : dict {mois: métriques}  si par_mois=True
    """
    df = pd.read_csv(csv_path)
    df["timestamp"] = pd.to_datetime(df["timestamp"], format="mixed", errors="coerce")

    colonnes_requises = [COL_REEL, COL_ESTIME, COL_ON_PRED, "timestamp"]
    manquantes = [c for c in colonnes_requises if c not in df.columns]
    if manquantes:
        raise ValueError(f"Colonnes manquantes dans {Path(csv_path).name} : {manquantes}")

    df = df.sort_values("timestamp").reset_index(drop=True)
    df["mois"]   = df["timestamp"].dt.month
    df["saison"] = df["mois"].map({
        12: "Hiver",     1: "Hiver",     2: "Hiver",
         3: "Printemps", 4: "Printemps", 5: "Printemps",
         6: "Été",       7: "Été",       8: "Été",
         9: "Automne",  10: "Automne",  11: "Automne",
    })

    # Extraire dataid/période/année depuis le nom du fichier
    # Format : resultats_<dataid>_<periode>_<annee>_complet.csv
    nom     = Path(csv_path).stem   # ex. resultats_3864_annee_2015_complet
    parties = nom.split("_")
    try:
        dataid  = parties[1]
        periode = parties[2]
        annee   = parties[3]
    except IndexError:
        dataid, periode, annee = "—", "—", "—"

    def _metriques_df(sous_df: pd.DataFrame) -> dict:
        """Calcule les métriques sur un sous-ensemble du DataFrame."""
        if sous_df.empty:
            return {}
        y_reel   = pd.to_numeric(sous_df[COL_REEL],    errors="coerce").fillna(0).to_numpy()
        y_estime = pd.to_numeric(sous_df[COL_ESTIME],  errors="coerce").fillna(0).to_numpy()
        o_pred   = pd.to_numeric(sous_df[COL_ON_PRED], errors="coerce").fillna(0).to_numpy()
        o_reel   = (y_reel > seuil_on).astype(int)

        rmse        = float(np.sqrt(np.mean((y_reel - y_estime) ** 2)))
        mae         = float(np.mean(np.abs(y_reel - y_estime)))
        energy_frac = float(y_estime.sum() / max(y_reel.sum(), 1e-9))
        norm_err    = float(np.abs(y_reel - y_estime).sum() / max(np.abs(y_reel).sum(), 1e-9))

        o_pred_bin = np.round(o_pred).astype(int)
        tp = int(((o_reel == 1) & (o_pred_bin == 1)).sum())
        tn = int(((o_reel == 0) & (o_pred_bin == 0)).sum())
        fp = int(((o_reel == 0) & (o_pred_bin == 1)).sum())
        fn = int(((o_reel == 1) & (o_pred_bin == 0)).sum())

        recall    = tp / (tp + fn) if (tp + fn) > 0 else float("nan")
        precision = tp / (tp + fp) if (tp + fp) > 0 else float("nan")
        f1 = (2 * precision * recall / (precision + recall)
              if not (np.isnan(precision) or np.isnan(recall))
              and (precision + recall) > 0 else float("nan"))
        fpr = fp / (fp + tn) if (fp + tn) > 0 else float("nan")

        return {
            "N pas"       : len(sous_df),
            "RMSE (kW)"   : round(rmse,        4),
            "MAE (kW)"    : round(mae,          4),
            "Energy Frac" : round(energy_frac,  4),
            "Norm Err"    : round(norm_err,      4),
            "Recall"      : round(recall,        4) if not np.isnan(recall)    else float("nan"),
            "Precision"   : round(precision,     4) if not np.isnan(precision) else float("nan"),
            "F1"          : round(f1,            4) if not np.isnan(f1)        else float("nan"),
            "FPR"         : round(fpr,           4) if not np.isnan(fpr)       else float("nan"),
            "TP": tp, "TN": tn, "FP": fp, "FN": fn,
        }

    resultats = {
        "fichier"           : Path(csv_path).name,
        "dataid"            : dataid,
        "periode"           : periode,
        "annee"             : annee,
        "metriques_globales": _metriques_df(df),
    }

    if par_saison:
        resultats["metriques_par_saison"] = {
            s: _metriques_df(df[df["saison"] == s])
            for s in ["Hiver", "Printemps", "Été", "Automne"]
            if (df["saison"] == s).any()
        }

    if par_mois:
        noms_mois = {1:"Jan", 2:"Fév", 3:"Mar", 4:"Avr", 5:"Mai", 6:"Jun",
                     7:"Jul", 8:"Aoû", 9:"Sep", 10:"Oct", 11:"Nov", 12:"Déc"}
        resultats["metriques_par_mois"] = {
            noms_mois[m]: _metriques_df(df[df["mois"] == m])
            for m in sorted(df["mois"].unique())
        }

    return resultats


def afficher_evaluation_periode(resultats: dict) -> None:
    """Affiche les métriques d'une évaluation de période de façon lisible."""

    print(f"\n{'=' * 65}")
    print(f"ÉVALUATION — Client {resultats['dataid']}  |  "
          f"{resultats['periode'].upper()} {resultats['annee']}")
    print(f"{'=' * 65}")

    g = resultats["metriques_globales"]
    print(f"\n  Métriques globales ({g.get('N pas', '?')} pas) :")
    for k, v in g.items():
        if k != "N pas":
            print(f"    {k:<15} : {v}")

    if "metriques_par_saison" in resultats:
        print(f"\n  Par saison :")
        df_s = pd.DataFrame(resultats["metriques_par_saison"]).T[
            ["N pas", "RMSE (kW)", "MAE (kW)", "Energy Frac", "F1"]
        ]
        print(df_s.to_string())

    if "metriques_par_mois" in resultats:
        print(f"\n  Par mois :")
        df_m = pd.DataFrame(resultats["metriques_par_mois"]).T[
            ["N pas", "RMSE (kW)", "MAE (kW)", "Energy Frac", "F1"]
        ]
        print(df_m.to_string())


def _cel(val) -> object:
    """Convertit toute valeur non-écrivable par openpyxl en '—'.
    Traite Python float nan, np.float64 nan, np.float32, inf, None, etc.
    """
    if val is None:
        return "—"
    if isinstance(val, str):
        return val
    # Convertir les scalaires numpy en types Python natifs
    if hasattr(val, "item"):
        val = val.item()
    # Maintenant val est un type Python natif (int, float, bool…)
    if isinstance(val, float) and (val != val or val == float("inf") or val == float("-inf")):
        return "—"
    return val


def sauvegarder_excel_periode(resultats: dict, out_path: str) -> None:
    """Alias conservé pour compatibilité — délègue à sauvegarder_excel_tous_clients."""
    sauvegarder_excel_tous_clients([resultats], out_path)


def sauvegarder_excel_tous_clients(tous_resultats: list[dict], out_path: str) -> None:
    """
    Génère un rapport Excel unique regroupant tous les clients.

    Onglets produits :
      • « Résumé »           — une ligne par client (métriques globales) + ligne de moyennes
      • « [dataid]_[annee] » — pour chaque client : global, par saison, par mois
    """
    wb = Workbook()

    # ── Palette & helpers ─────────────────────────────────────────────
    BLUE_DARK  = "1F3864"
    BLUE_MID   = "2E75B6"
    BLUE_LIGHT = "D6E4F0"
    TEAL_DARK  = "1B6B5A"
    TEAL_MID   = "2E8B6F"
    WHITE      = "FFFFFF"
    GREY       = "F2F2F2"

    def fill(c):
        return PatternFill("solid", start_color=c)
    def font_(bold=False, color="000000", size=10):
        return Font(name="Arial", bold=bold, color=color, size=size)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin   = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    COLS_DETAIL = [
        "N pas", "RMSE (kW)", "MAE (kW)", "Energy Frac", "Norm Err",
        "Recall", "Precision", "F1", "FPR", "TP", "TN", "FP", "FN",
    ]

    # ── Onglet Résumé ─────────────────────────────────────────────────
    ws_res = wb.active
    ws_res.title = "Résumé"
    ws_res.sheet_view.showGridLines = False

    COLS_RESUME = [
        "Client", "Période", "Année", "N pas",
        "RMSE (kW)", "MAE (kW)", "Energy Frac", "Norm Err",
        "Recall", "Precision", "F1", "FPR",
        "TP", "TN", "FP", "FN",
    ]
    n_cols_res = len(COLS_RESUME)

    # Titre
    ws_res.merge_cells(f"A1:{get_column_letter(n_cols_res)}1")
    ws_res["A1"] = "THERMO — Récapitulatif des métriques (tous clients)"
    ws_res["A1"].font      = font_(bold=True, color=WHITE, size=13)
    ws_res["A1"].fill      = fill(BLUE_DARK)
    ws_res["A1"].alignment = center
    ws_res.row_dimensions[1].height = 28

    # En-têtes
    ws_res.row_dimensions[2].height = 32
    for col_idx, h in enumerate(COLS_RESUME, start=1):
        c = ws_res.cell(row=2, column=col_idx, value=h)
        c.font      = font_(bold=True, color=WHITE, size=10)
        c.fill      = fill(BLUE_MID)
        c.alignment = center
        c.border    = border

    # Données — une ligne par client
    accumulated = {col: [] for col in COLS_RESUME[3:]}
    for i, res in enumerate(tous_resultats, start=3):
        g     = res["metriques_globales"]
        shade = GREY if i % 2 == 0 else WHITE
        ws_res.row_dimensions[i].height = 18
        row_vals = [
            res["dataid"], res["periode"].upper(), res["annee"],
            g.get("N pas", "—"),
            g.get("RMSE (kW)", "—"), g.get("MAE (kW)", "—"),
            g.get("Energy Frac", "—"), g.get("Norm Err", "—"),
            g.get("Recall", "—"), g.get("Precision", "—"),
            g.get("F1", "—"), g.get("FPR", "—"),
            g.get("TP", "—"), g.get("TN", "—"),
            g.get("FP", "—"), g.get("FN", "—"),
        ]
        for col_idx, val in enumerate(row_vals, start=1):
            c = ws_res.cell(row=i, column=col_idx, value=_cel(val))
            c.fill      = fill(shade)
            c.border    = border
            c.font      = font_(size=10)
            c.alignment = left if col_idx == 1 else center
        for col_name, val in zip(COLS_RESUME[3:], row_vals[3:]):
            if isinstance(val, (int, float)) and not (val != val):
                accumulated[col_name].append(float(val))

    # Ligne de moyennes
    row_moy = len(tous_resultats) + 3
    ws_res.row_dimensions[row_moy].height = 20
    ws_res.merge_cells(f"A{row_moy}:C{row_moy}")
    ws_res.cell(row=row_moy, column=1, value="MOYENNE").font      = font_(bold=True, color=WHITE, size=10)
    ws_res.cell(row=row_moy, column=1).fill      = fill(BLUE_DARK)
    ws_res.cell(row=row_moy, column=1).alignment = center
    ws_res.cell(row=row_moy, column=1).border    = border
    for col_idx, col_name in enumerate(COLS_RESUME[3:], start=4):
        vals = accumulated[col_name]
        val  = round(float(np.mean(vals)), 4) if vals else "—"
        c = ws_res.cell(row=row_moy, column=col_idx, value=_cel(val))
        c.fill      = fill(BLUE_LIGHT)
        c.border    = border
        c.font      = font_(bold=True, size=10)
        c.alignment = center

    col_widths_res = [12, 10, 8, 8] + [11] * (n_cols_res - 4)
    for i, w in enumerate(col_widths_res, start=1):
        ws_res.column_dimensions[get_column_letter(i)].width = w
    ws_res.freeze_panes = "A3"

    # ── Onglets détaillés — un par client ─────────────────────────────
    def _ecrire_section(ws, row_start: int, titre_section: str,
                        couleur_titre: str, donnees: dict) -> int:
        """Écrit un bloc titre+en-têtes+lignes. Retourne la prochaine ligne libre."""
        n_cols = 1 + len(COLS_DETAIL)
        ws.merge_cells(f"A{row_start}:{get_column_letter(n_cols)}{row_start}")
        ws.cell(row=row_start, column=1, value=titre_section).font      = font_(bold=True, color=WHITE, size=11)
        ws.cell(row=row_start, column=1).fill      = fill(couleur_titre)
        ws.cell(row=row_start, column=1).alignment = center
        ws.row_dimensions[row_start].height = 22
        row_start += 1

        en_tetes = ["Groupe"] + COLS_DETAIL
        for col_idx, h in enumerate(en_tetes, start=1):
            c = ws.cell(row=row_start, column=col_idx, value=h)
            c.font      = font_(bold=True, color=WHITE, size=10)
            c.fill      = fill(BLUE_MID)
            c.alignment = center
            c.border    = border
        ws.row_dimensions[row_start].height = 26
        row_start += 1

        for j, (label, metriques) in enumerate(donnees.items()):
            shade = GREY if j % 2 == 0 else WHITE
            ws.row_dimensions[row_start].height = 18
            ws.cell(row=row_start, column=1, value=label).font      = font_(bold=True, size=10)
            ws.cell(row=row_start, column=1).fill      = fill(shade)
            ws.cell(row=row_start, column=1).alignment = center
            ws.cell(row=row_start, column=1).border    = border
            for col_idx, col in enumerate(COLS_DETAIL, start=2):
                c = ws.cell(row=row_start, column=col_idx,
                            value=_cel(metriques.get(col, "—")))
                c.fill      = fill(shade)
                c.border    = border
                c.font      = font_(size=10)
                c.alignment = center
            row_start += 1

        return row_start + 1   # ligne vide de séparation

    for res in tous_resultats:
        onglet_nom = f"{res['dataid']}_{res['annee']}"[:31]
        ws = wb.create_sheet(onglet_nom)
        ws.sheet_view.showGridLines = False

        n_cols = 1 + len(COLS_DETAIL)
        ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
        ws["A1"] = (f"Client {res['dataid']} — "
                    f"{res['periode'].upper()} {res['annee']}")
        ws["A1"].font      = font_(bold=True, color=WHITE, size=13)
        ws["A1"].fill      = fill(BLUE_DARK)
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 28

        next_row = 3

        next_row = _ecrire_section(
            ws, next_row,
            f"Métriques globales ({res['metriques_globales'].get('N pas', '?')}{chr(160)}pas)",
            BLUE_MID,
            {"Global": res["metriques_globales"]},
        )

        if res.get("metriques_par_saison"):
            next_row = _ecrire_section(
                ws, next_row, "Par saison", TEAL_MID, res["metriques_par_saison"])

        if res.get("metriques_par_mois"):
            next_row = _ecrire_section(
                ws, next_row, "Par mois", TEAL_DARK, res["metriques_par_mois"])

        for col_idx, w in enumerate([18] + [12] * len(COLS_DETAIL), start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w
        ws.freeze_panes = "A3"

    wb.save(out_path)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Calcul MAE / RMSE à partir des CSV de désagrégation",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemples :
  python src/calcul_metriques.py
  python src/calcul_metriques.py --dossier output/california
  python src/calcul_metriques.py --rapport output/metriques_final.xlsx
  python src/calcul_metriques.py --seuil_on 0.10
        """,
    )
    parser.add_argument(
        "--dossier", type=str, default=str(OUTPUT_DIR),
        help="Dossier contenant les CSV de désagrégation (défaut : output/)",
    )
    parser.add_argument(
        "--rapport", type=str,
        default=str(OUTPUT_DIR / "metriques_desagregation.xlsx"),
        help="Chemin du rapport Excel de sortie (défaut : output/metriques_desagregation.xlsx)",
    )
    parser.add_argument(
        "--seuil_on", type=float, default=0.05,
        help="Seuil de binarisation P_reel_clim → ON en kW (défaut : 0.05)",
    )
    parser.add_argument(
        "--periode", type=str, default=None,
        help=(
            "Si fourni, évalue les fichiers '*_complet.csv' au lieu des "
            "fichiers hebdomadaires. Ex : --periode annee ou --periode ete"
        ),
    )
    args = parser.parse_args()

    Path(args.rapport).parent.mkdir(parents=True, exist_ok=True)

    print("=" * 60)
    print("CALCUL DES MÉTRIQUES — THERMO NILM")
    print("=" * 60)
    print(f"  Dossier CSV  : {args.dossier}")
    print(f"  Rapport Excel: {args.rapport}")
    print(f"  Seuil ON     : {args.seuil_on} kW")

    # ── Mode période : fichiers *_complet.csv ────────────────────────────
    if args.periode is not None:
        pattern_complet = str(
            Path(args.dossier) / f"resultats_*_{args.periode}_*_complet.csv"
        )
        csv_complets = sorted(glob.glob(pattern_complet))

        if not csv_complets:
            print(f"\n⚠️  Aucun fichier '*_{args.periode}_*_complet.csv' trouvé dans : {args.dossier}")
            print("   Lancez d'abord optiPeriode.py pour générer les CSV consolidés.")
            return

        print(f"\n  {len(csv_complets)} fichier(s) complet(s) trouvé(s)\n")

        tous_resultats = []
        for csv_path in csv_complets:
            print(f"  {Path(csv_path).name}")
            try:
                res = evaluer_periode(csv_path, seuil_on=args.seuil_on)
                afficher_evaluation_periode(res)
                tous_resultats.append(res)
                print(f"  ✔ OK — Client {res['dataid']} ({res['periode'].upper()} {res['annee']})")
            except Exception as e:
                print(f"  ✗ Erreur : {e}")

        if not tous_resultats:
            print("\n⚠️  Aucun résultat valide — rapport non généré.")
            return

        sauvegarder_excel_tous_clients(tous_resultats, args.rapport)
        print(f"\n✔ Rapport unique sauvegardé : {args.rapport}")
        print(f"   Onglets : Résumé + {len(tous_resultats)} client(s)")
        return

    # ── Mode hebdomadaire : fichiers resultats_desagregation_*.csv ───────
    pattern   = str(Path(args.dossier) / "resultats_desagregation_*.csv")
    csv_files = sorted(glob.glob(pattern))

    if not csv_files:
        print(f"\n⚠️  Aucun fichier 'resultats_desagregation_*.csv' trouvé dans : {args.dossier}")
        print("   Lancez d'abord optiUneSemaine.py pour générer les CSV.")
        return

    print(f"\n  {len(csv_files)} fichier(s) trouvé(s)\n")

    resultats = []
    for csv_path in csv_files:
        nom = Path(csv_path).name
        print(f"  {nom} ... ", end="", flush=True)
        try:
            metriques = calculer_metriques(csv_path, seuil_on=args.seuil_on)
            resultats.append(metriques)
            print(f"RMSE={metriques['RMSE (kW)']:.4f}  MAE={metriques['MAE (kW)']:.4f}  F1={metriques['F1']}")
        except Exception as e:
            print(f"ERREUR : {e}")

    if not resultats:
        print("\n⚠️  Aucun résultat valide — rapport non généré.")
        return

    # ── Affichage du tableau récapitulatif ────────────────────────────
    print("\n" + "=" * 60)
    print("RÉCAPITULATIF")
    print("=" * 60)
    df_res = pd.DataFrame(resultats)[[
        "Client (dataid)", "Date début",
        "RMSE (kW)", "MAE (kW)", "Energy Frac", "Norm Err",
        "Recall", "F1", "FPR", "Gap réel (%)",   # ← ajouter
    ]]
    print(df_res.to_string(index=False))

    # ── Moyennes ──────────────────────────────────────────────────────
    print("\n  Moyennes :")
    for col in ["RMSE (kW)", "MAE (kW)", "Energy Frac", "Norm Err", "Recall", "F1", "FPR", "Gap réel (%)"]:
        vals = pd.to_numeric(df_res[col], errors="coerce").dropna()
        if len(vals) > 0:
            print(f"    {col:<15} : {vals.mean():.4f}")

    # ── Sauvegarde Excel ──────────────────────────────────────────────
    sauvegarder_excel(resultats, args.rapport)
    print(f"\n✔ Rapport sauvegardé : {args.rapport}")


if __name__ == "__main__":
    main()