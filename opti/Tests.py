#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Procédure de test — Projet THERMO NILM.

Lit les CSV de désagrégation produits par optiUneSemaine.py depuis output/
et génère le rapport Excel dans output/rapport_tests.xlsx.

Usage :
    python src/Tests.py
    python src/Tests.py --dossier output/mes_resultats
    python src/Tests.py --dossier output --rapport output/mon_rapport.xlsx
"""

from __future__ import annotations

import argparse
import os
import glob
import tempfile
from pathlib import Path
from typing import Any, Dict, Optional, Sequence

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════
# CONFIGURATION — chemins relatifs au projet
# ══════════════════════════════════════════════════════════════════════

PROJECT_DIR = Path(__file__).resolve().parent.parent
OUTPUT_DIR  = PROJECT_DIR / "output"

# Colonnes attendues dans les CSV produits par optiUneSemaine.py
COL_REEL_CLIM   = "P_reel_clim"        # puissance réelle climatisation
COL_ESTIME_CLIM = "P_estime_clim"      # puissance estimée climatisation
COL_ON_PRED     = "o_climatisation"    # état ON/OFF prédit (binaire)

THRESHOLDS_DEFAUT: Dict[str, Any] = {
    "recall_min"            : 0.70,
    "f1_min"                : 0.70,
    "fpr_max"               : 0.15,
    "energy_frac_min"       : 0.75,
    "energy_frac_max"       : 1.25,
    "norm_err_max"          : 0.45,
    "binarisation_threshold": 0.15,
}


# ══════════════════════════════════════════════════════════════════════
# UTILITAIRES
# ══════════════════════════════════════════════════════════════════════

def load_results_csv(
    csv_path: str,
    encoding: str = "utf-8",
    sep: Optional[str] = None,
) -> pd.DataFrame:
    if sep is None:
        return pd.read_csv(csv_path, encoding=encoding, sep=None, engine="python")
    return pd.read_csv(csv_path, encoding=encoding, sep=sep)


def _require_cols(df: pd.DataFrame, cols: Sequence[str]) -> None:
    missing = [c for c in cols if c not in df.columns]
    if missing:
        raise ValueError(f"Colonnes manquantes dans le CSV : {missing}")


def _to_binary(x: pd.Series) -> np.ndarray:
    v = pd.to_numeric(x, errors="coerce").fillna(0)
    return (v != 0).astype(int).to_numpy()


def _result(
    test_id: str,
    passed: bool,
    metrics: Dict[str, float],
    details: Dict[str, Any],
) -> Dict[str, Any]:
    return {"test_id": test_id, "passed": passed, "metrics": metrics, "details": details}


# ══════════════════════════════════════════════════════════════════════
# MÉTRIQUES
# ══════════════════════════════════════════════════════════════════════

def confusion_binary(
    y_true: np.ndarray,
    y_pred: np.ndarray,
) -> Dict[str, int]:
    yt = (np.asarray(y_true).astype(int) != 0).astype(int)
    yp = (np.asarray(y_pred).astype(int) != 0).astype(int)
    return {
        "tp": int(((yt == 1) & (yp == 1)).sum()),
        "tn": int(((yt == 0) & (yp == 0)).sum()),
        "fp": int(((yt == 0) & (yp == 1)).sum()),
        "fn": int(((yt == 1) & (yp == 0)).sum()),
    }


def detection_metrics(
    y_true: np.ndarray,
    y_pred: np.ndarray,
) -> Dict[str, float]:
    c  = confusion_binary(y_true, y_pred)
    tp, tn, fp, fn = c["tp"], c["tn"], c["fp"], c["fn"]
    recall    = tp / (tp + fn) if (tp + fn) else np.nan
    precision = tp / (tp + fp) if (tp + fp) else np.nan
    f1        = (2 * precision * recall / (precision + recall)
                 if (not np.isnan(precision) and not np.isnan(recall)
                     and (precision + recall) > 0)
                 else np.nan)
    fpr       = fp / (fp + tn) if (fp + tn) else np.nan
    return {
        "recall"   : float(recall),
        "precision": float(precision),
        "f1"       : float(f1),
        "fpr"      : float(fpr),
    }


def energy_fraction(
    y_true: np.ndarray,
    y_pred: np.ndarray,
    eps: float = 1e-9,
) -> float:
    yt = np.asarray(y_true, dtype=float)
    yp = np.asarray(y_pred, dtype=float)
    return float(yp.sum() / max(yt.sum(), eps))


def norm_l1_error(
    y_true: np.ndarray,
    y_pred: np.ndarray,
    eps: float = 1e-9,
) -> float:
    yt = np.asarray(y_true, dtype=float)
    yp = np.asarray(y_pred, dtype=float)
    return float(np.abs(yp - yt).sum() / max(np.abs(yt).sum(), eps))


# ══════════════════════════════════════════════════════════════════════
# TESTS INDIVIDUELS
# ══════════════════════════════════════════════════════════════════════

def test_T02_detection(
    csv_path: str,
    *,
    true_col: str,
    pred_col: str,
    filter_expr: Optional[str] = None,
    recall_min: float = 0.70,
    f1_min: float = 0.70,
) -> Dict[str, Any]:
    df = load_results_csv(csv_path)
    _require_cols(df, [true_col, pred_col])
    if filter_expr:
        df = df.query(filter_expr).copy()
    y_true = _to_binary(df[true_col])
    y_pred = _to_binary(df[pred_col])
    met    = detection_metrics(y_true, y_pred)
    conf   = confusion_binary(y_true, y_pred)
    passed = (
        not np.isnan(met["recall"]) and not np.isnan(met["f1"])
        and met["recall"] >= recall_min
        and met["f1"]     >= f1_min
    )
    return _result("T-02", passed, met, {
        "confusion": conf, "n": int(len(df)), "filter": filter_expr,
        "targets": {"recall_min": recall_min, "f1_min": f1_min},
    })


def test_T03_false_positives(
    csv_path: str,
    *,
    true_col: str,
    pred_col: str,
    filter_expr: Optional[str] = None,
    fpr_max: float = 0.15,
) -> Dict[str, Any]:
    df = load_results_csv(csv_path)
    _require_cols(df, [true_col, pred_col])
    if filter_expr:
        df = df.query(filter_expr).copy()
    y_true = _to_binary(df[true_col])
    y_pred = _to_binary(df[pred_col])
    met    = detection_metrics(y_true, y_pred)
    conf   = confusion_binary(y_true, y_pred)
    passed = (not np.isnan(met["fpr"])) and (met["fpr"] <= fpr_max)
    return _result("T-03", passed, {"fpr": met["fpr"]}, {
        "confusion": conf, "n": int(len(df)), "filter": filter_expr,
        "target": {"fpr_max": fpr_max},
    })


def test_T05_prediction(
    csv_path: str,
    *,
    true_power_col: str,
    pred_power_col: str,
    filter_expr: Optional[str] = None,
    energy_frac_min: float = 0.75,
    energy_frac_max: float = 1.25,
    norm_err_max: float = 0.45,
) -> Dict[str, Any]:
    df = load_results_csv(csv_path)
    _require_cols(df, [true_power_col, pred_power_col])
    if filter_expr:
        df = df.query(filter_expr).copy()
    yt  = pd.to_numeric(df[true_power_col],  errors="coerce").fillna(0).to_numpy(dtype=float)
    yp  = pd.to_numeric(df[pred_power_col], errors="coerce").fillna(0).to_numpy(dtype=float)
    met = {
        "energy_frac": energy_fraction(yt, yp),
        "norm_err"   : norm_l1_error(yt, yp),
    }
    passed = (
        energy_frac_min <= met["energy_frac"] <= energy_frac_max
        and met["norm_err"] <= norm_err_max
    )
    return _result("T-05", passed, met, {
        "n": int(len(df)), "filter": filter_expr,
        "targets": {
            "energy_frac_min": energy_frac_min,
            "energy_frac_max": energy_frac_max,
            "norm_err_max"   : norm_err_max,
        },
    })


# ══════════════════════════════════════════════════════════════════════
# PIPELINE PAR FICHIER
# ══════════════════════════════════════════════════════════════════════

def run_tests_on_file(
    csv_path: str,
    thresholds: Dict[str, Any],
) -> Dict[str, Any]:
    """
    Exécute T-02, T-03, T-05 sur un CSV de désagrégation produit par
    optiUneSemaine.py.

    Colonnes attendues dans le CSV :
      - P_reel_clim       : puissance réelle climatisation (kW)
      - P_estime_clim     : puissance estimée par l'optimiseur (kW)
      - o_climatisation   : état ON/OFF prédit (0 ou 1)
    """
    df    = load_results_csv(csv_path)
    seuil = thresholds.get("binarisation_threshold", 0.0)

    # Vérifier les colonnes attendues
    _require_cols(df, [COL_REEL_CLIM, COL_ESTIME_CLIM, COL_ON_PRED])

    # Dériver la vérité terrain binaire à partir de la puissance réelle
    df["ac_on_true"] = (
        pd.to_numeric(df[COL_REEL_CLIM], errors="coerce").fillna(0) > seuil
    ).astype(int)

    # Fichier temporaire dans le dossier système — ne pollue pas output/
    with tempfile.NamedTemporaryFile(
        mode="w", suffix=".csv", delete=False, encoding="utf-8"
    ) as tmp_f:
        tmp_path = tmp_f.name
    df.to_csv(tmp_path, index=False)

    try:
        r_T02 = test_T02_detection(
            csv_path =tmp_path,
            true_col ="ac_on_true",
            pred_col =COL_ON_PRED,
            recall_min=thresholds["recall_min"],
            f1_min    =thresholds["f1_min"],
        )
        r_T03 = test_T03_false_positives(
            csv_path=tmp_path,
            true_col=    "ac_on_true",
            pred_col=COL_ON_PRED,
            fpr_max =thresholds["fpr_max"],
        )
        r_T05 = test_T05_prediction(
            csv_path      =tmp_path,
            true_power_col=COL_REEL_CLIM,
            pred_power_col=COL_ESTIME_CLIM,
            energy_frac_min=thresholds["energy_frac_min"],
            energy_frac_max=thresholds["energy_frac_max"],
            norm_err_max   =thresholds["norm_err_max"],
        )
    finally:
        os.remove(tmp_path)

    return {"T02": r_T02, "T03": r_T03, "T05": r_T05}


# ══════════════════════════════════════════════════════════════════════
# RAPPORT EXCEL
# ══════════════════════════════════════════════════════════════════════

def save_excel_report(
    all_results: list[Dict[str, Any]],
    out_path: str,
    thresholds: Dict[str, Any],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport de tests"

    # ── Styles ─────────────────────────────────────────────────────────
    header_font    = Font(name="Arial", bold=True, color="FFFFFF",  size=11)
    header_fill    = PatternFill("solid", start_color="2F5496")
    subheader_fill = PatternFill("solid", start_color="D9E1F2")
    pass_fill      = PatternFill("solid", start_color="C6EFCE")
    fail_fill      = PatternFill("solid", start_color="FFC7CE")
    pass_font      = Font(name="Arial", bold=True, color="276221")
    fail_font      = Font(name="Arial", bold=True, color="9C0006")
    normal_font    = Font(name="Arial", size=10)
    center         = Alignment(horizontal="center", vertical="center")
    thin           = Side(style="thin", color="BFBFBF")
    border         = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Fichier CSV", "N",
        "T-02 Résultat", "Recall", "Precision", "F1",
        "T-03 Résultat", "FPR",
        "T-05 Résultat", "Energy Frac", "Norm Err",
        "Résultat global",
    ]
    seuils_row = [
        "— Seuils —", "—",
        "—", f"≥ {thresholds['recall_min']:.0%}", "—", f"≥ {thresholds['f1_min']:.0%}",
        "—", f"≤ {thresholds['fpr_max']:.0%}",
        "—",
        f"{thresholds['energy_frac_min']:.0%} – {thresholds['energy_frac_max']:.0%}",
        f"≤ {thresholds['norm_err_max']:.0%}",
        "—",
    ]
    descriptions = {
        "T-02": (
            "Mesure la capacité de l'algorithme à détecter correctement les instants où la "
            "climatisation est allumée (ON), en évaluant le rappel (proportion de vrais ON "
            "détectés) et le score F1 (équilibre entre détection et précision)."
        ),
        "T-03": (
            "Mesure la proportion de pas de temps où la climatisation est réellement éteinte "
            "(OFF) mais que l'algorithme prédit à tort comme allumée (faux positifs)."
        ),
        "T-05": (
            "Mesure la qualité de l'estimation de la puissance consommée par la climatisation : "
            "la fraction d'énergie évalue si l'énergie totale estimée correspond à la réalité, "
            "et l'erreur normalisée quantifie l'écart cumulé entre la puissance estimée et la "
            "puissance réelle."
        ),
    }

    # Ligne 1 — Titre
    ws.merge_cells("A1:L1")
    ws["A1"] = (
        f"Rapport de tests NILM — Climatisation  |  "
        f"Seuil ON : P_reel_clim > {thresholds['binarisation_threshold']} kW"
    )
    ws["A1"].font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", start_color="1F3864")
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    # Ligne 2 — En-têtes
    for col_idx, h in enumerate(headers, start=1):
        cell           = ws.cell(row=2, column=col_idx, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border

    # Ligne 3 — Seuils
    for col_idx, s in enumerate(seuils_row, start=1):
        cell           = ws.cell(row=3, column=col_idx, value=s)
        cell.font      = Font(name="Arial", bold=True, size=9, color="2F5496")
        cell.fill      = subheader_fill
        cell.alignment = center
        cell.border    = border

    # Lignes de données
    for entry in all_results:
        fname   = entry["filename"]
        T02     = entry["results"]["T02"]
        T03     = entry["results"]["T03"]
        T05     = entry["results"]["T05"]
        n       = T02["details"]["n"]
        overall = T02["passed"] and T03["passed"] and T05["passed"]

        row = [
            fname, n,
            "PASS" if T02["passed"] else "FAIL",
            round(T02["metrics"].get("recall",      float("nan")), 4),
            round(T02["metrics"].get("precision",   float("nan")), 4),
            round(T02["metrics"].get("f1",          float("nan")), 4),
            "PASS" if T03["passed"] else "FAIL",
            round(T03["metrics"].get("fpr",         float("nan")), 4),
            "PASS" if T05["passed"] else "FAIL",
            round(T05["metrics"].get("energy_frac", float("nan")), 4),
            round(T05["metrics"].get("norm_err",    float("nan")), 4),
            "PASS" if overall else "FAIL",
        ]
        ws.append(row)
        r = ws.max_row

        for col_idx in range(1, 13):
            cell           = ws.cell(row=r, column=col_idx)
            cell.font      = normal_font
            cell.alignment = center
            cell.border    = border

        for col_idx in [3, 7, 9, 12]:
            cell = ws.cell(row=r, column=col_idx)
            if cell.value == "PASS":
                cell.fill = pass_fill
                cell.font = pass_font
            else:
                cell.fill = fail_fill
                cell.font = fail_font

    # Largeurs de colonnes
    col_widths = [40, 8, 14, 10, 12, 10, 14, 10, 14, 13, 12, 16]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

    # Feuille 2 — Descriptions des tests
    ws2 = wb.create_sheet(title="Description des tests")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 100

    ws2.merge_cells("A1:B1")
    ws2["A1"] = "Description des tests"
    ws2["A1"].font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws2["A1"].fill      = PatternFill("solid", start_color="1F3864")
    ws2["A1"].alignment = center
    ws2.row_dimensions[1].height = 28

    for col_idx, h in enumerate(["Test", "Description"], start=1):
        cell           = ws2.cell(row=2, column=col_idx, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border

    for row_idx, (test_id, desc) in enumerate(
        [("T-02", descriptions["T-02"]),
         ("T-03", descriptions["T-03"]),
         ("T-05", descriptions["T-05"])],
        start=3,
    ):
        cell_id           = ws2.cell(row=row_idx, column=1, value=test_id)
        cell_id.font      = Font(name="Arial", bold=True, size=10)
        cell_id.alignment = center
        cell_id.border    = border

        cell_desc           = ws2.cell(row=row_idx, column=2, value=desc)
        cell_desc.font      = Font(name="Arial", size=10)
        cell_desc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell_desc.border    = border
        ws2.row_dimensions[row_idx].height = 60

    wb.save(out_path)


# ══════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Tests NILM — lit les CSV de désagrégation depuis output/",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemples :
  python src/Tests.py
  python src/Tests.py --dossier output/mes_resultats
  python src/Tests.py --rapport output/rapport_final.xlsx
  python src/Tests.py --recall_min 0.60 --f1_min 0.60
        """,
    )
    parser.add_argument(
        "--dossier", type=str,
        default=str(OUTPUT_DIR),
        help=(
            "Dossier contenant les CSV de désagrégation "
            f"(défaut : output/)"
        ),
    )
    parser.add_argument(
        "--rapport", type=str,
        default=str(OUTPUT_DIR / "rapport_tests.xlsx"),
        help="Chemin du rapport Excel de sortie (défaut : output/rapport_tests.xlsx)",
    )
    # Seuils modifiables en CLI
    parser.add_argument("--recall_min",      type=float, default=THRESHOLDS_DEFAUT["recall_min"])
    parser.add_argument("--f1_min",          type=float, default=THRESHOLDS_DEFAUT["f1_min"])
    parser.add_argument("--fpr_max",         type=float, default=THRESHOLDS_DEFAUT["fpr_max"])
    parser.add_argument("--energy_frac_min", type=float, default=THRESHOLDS_DEFAUT["energy_frac_min"])
    parser.add_argument("--energy_frac_max", type=float, default=THRESHOLDS_DEFAUT["energy_frac_max"])
    parser.add_argument("--norm_err_max",    type=float, default=THRESHOLDS_DEFAUT["norm_err_max"])
    parser.add_argument("--seuil_on",        type=float, default=THRESHOLDS_DEFAUT["binarisation_threshold"],
                        help="Seuil de binarisation P_reel_clim → ON (kW)")
    args = parser.parse_args()

    thresholds = {
        "recall_min"            : args.recall_min,
        "f1_min"                : args.f1_min,
        "fpr_max"               : args.fpr_max,
        "energy_frac_min"       : args.energy_frac_min,
        "energy_frac_max"       : args.energy_frac_max,
        "norm_err_max"          : args.norm_err_max,
        "binarisation_threshold": args.seuil_on,
    }

    dossier_csv = args.dossier
    rapport_out = args.rapport

    # S'assurer que le dossier de sortie du rapport existe
    Path(rapport_out).parent.mkdir(parents=True, exist_ok=True)

    print("=" * 60)
    print("TESTS NILM — THERMO")
    print("=" * 60)
    print(f"  Dossier CSV  : {dossier_csv}")
    print(f"  Rapport Excel: {rapport_out}")
    print(f"  Colonnes     : réel={COL_REEL_CLIM}  estimé={COL_ESTIME_CLIM}  ON={COL_ON_PRED}")
    print(f"  Seuil ON     : P_reel_clim > {thresholds['binarisation_threshold']} kW")

    # Chercher uniquement les CSV de désagrégation (pas les tmp ni autres)
    pattern    = os.path.join(dossier_csv, "resultats_desagregation_*.csv")
    csv_files  = sorted(glob.glob(pattern))

    if not csv_files:
        print(f"\n⚠️  Aucun fichier 'resultats_desagregation_*.csv' trouvé dans : {dossier_csv}")
        print("   Lancez d'abord optiUneSemaine.py pour générer les CSV de désagrégation.")
        return

    print(f"\n  {len(csv_files)} fichier(s) trouvé(s)\n")

    all_results = []
    for csv_path in csv_files:
        fname = os.path.basename(csv_path)
        print(f"  Traitement : {fname} ... ", end="", flush=True)
        try:
            results = run_tests_on_file(csv_path, thresholds)
            all_results.append({"filename": fname, "results": results})
            overall = all(r["passed"] for r in results.values())
            print("PASS" if overall else "FAIL")
        except Exception as e:
            print(f"ERREUR : {e}")

    if all_results:
        save_excel_report(all_results, rapport_out, thresholds)
        print(f"\n✔ Rapport sauvegardé : {rapport_out}")
    else:
        print("\n⚠️  Aucun résultat valide — rapport non généré.")


if __name__ == "__main__":
    main()