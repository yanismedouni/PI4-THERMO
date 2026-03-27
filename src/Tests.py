#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Procédure de test.

Created on Fri Feb 26 2026
@author: catherinehenri
"""

from __future__ import annotations

import os
import glob
from typing import Any, Dict, Optional, Sequence

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def load_results_csv(
    csv_path: str,
    encoding: str = "utf-8",
    sep: Optional[str] = None,
) -> pd.DataFrame:
    if sep is None:
        return pd.read_csv(csv_path, encoding=encoding, sep=None, engine="python")
    return pd.read_csv(csv_path, encoding=encoding, sep=sep)


def _require_cols(df: pd.DataFrame, cols: Sequence[str]) -> None:
    missing = [c for c in cols if c not in df.columns]
    if missing:
        raise ValueError(f"Colonnes manquantes: {missing}")


def _to_binary(x: pd.Series) -> np.ndarray:
    v = pd.to_numeric(x, errors="coerce").fillna(0)
    return (v != 0).astype(int).to_numpy()


def confusion_binary(y_true: np.ndarray, y_pred: np.ndarray) -> Dict[str, int]:
    yt = (np.asarray(y_true).astype(int) != 0).astype(int)
    yp = (np.asarray(y_pred).astype(int) != 0).astype(int)
    tp = int(((yt == 1) & (yp == 1)).sum())
    tn = int(((yt == 0) & (yp == 0)).sum())
    fp = int(((yt == 0) & (yp == 1)).sum())
    fn = int(((yt == 1) & (yp == 0)).sum())
    return {"tp": tp, "tn": tn, "fp": fp, "fn": fn}


def detection_metrics(y_true: np.ndarray, y_pred: np.ndarray) -> Dict[str, float]:
    c = confusion_binary(y_true, y_pred)
    tp, tn, fp, fn = c["tp"], c["tn"], c["fp"], c["fn"]
    recall    = tp / (tp + fn) if (tp + fn) else np.nan
    precision = tp / (tp + fp) if (tp + fp) else np.nan
    f1        = (2 * precision * recall / (precision + recall)) if (precision + recall) else np.nan
    fpr       = fp / (fp + tn) if (fp + tn) else np.nan
    return {"recall": float(recall), "precision": float(precision), "f1": float(f1), "fpr": float(fpr)}


def energy_fraction(y_true: np.ndarray, y_pred: np.ndarray, eps: float = 1e-9) -> float:
    yt = np.asarray(y_true, dtype=float)
    yp = np.asarray(y_pred, dtype=float)
    return float(yp.sum() / max(yt.sum(), eps))


def norm_l1_error(y_true: np.ndarray, y_pred: np.ndarray, eps: float = 1e-9) -> float:
    yt = np.asarray(y_true, dtype=float)
    yp = np.asarray(y_pred, dtype=float)
    return float(np.abs(yp - yt).sum() / max(np.abs(yt).sum(), eps))


def _result(test_id: str, passed: bool, metrics: Dict[str, float], details: Dict[str, Any]) -> Dict[str, Any]:
    return {"test_id": test_id, "passed": passed, "metrics": metrics, "details": details}


def test_T02_detection(
    csv_path: str,
    *,
    true_col: str,
    pred_col: str,
    filter_expr: Optional[str] = None,
    recall_min: float = 0.70,
    f1_min: float = 0.70,
) -> Dict[str, Any]:
    df = load_results_csv(csv_path)
    _require_cols(df, [true_col, pred_col])
    if filter_expr:
        df = df.query(filter_expr).copy()
    y_true = _to_binary(df[true_col])
    y_pred = _to_binary(df[pred_col])
    met = detection_metrics(y_true, y_pred)
    conf = confusion_binary(y_true, y_pred)
    passed = (
        not np.isnan(met["recall"])
        and not np.isnan(met["f1"])
        and met["recall"] >= recall_min
        and met["f1"] >= f1_min
    )
    return _result("T-02", passed, met,
        {"confusion": conf, "n": int(len(df)), "filter": filter_expr,
         "targets": {"recall_min": recall_min, "f1_min": f1_min}})


def test_T03_false_positives(
    csv_path: str,
    *,
    true_col: str,
    pred_col: str,
    filter_expr: Optional[str] = None,
    fpr_max: float = 0.15,
) -> Dict[str, Any]:
    df = load_results_csv(csv_path)
    _require_cols(df, [true_col, pred_col])
    if filter_expr:
        df = df.query(filter_expr).copy()
    y_true = _to_binary(df[true_col])
    y_pred = _to_binary(df[pred_col])
    met = detection_metrics(y_true, y_pred)
    conf = confusion_binary(y_true, y_pred)
    passed = (not np.isnan(met["fpr"])) and (met["fpr"] <= fpr_max)
    return _result("T-03", passed, {"fpr": met["fpr"]},
        {"confusion": conf, "n": int(len(df)), "filter": filter_expr,
         "target": {"fpr_max": fpr_max}})


def test_T04_fpr_three_regions(
    csv_path: str,
    *,
    region_col: str,
    true_col: str,
    pred_col: str,
    regions: Sequence[str],
    base_filter_expr: Optional[str] = None,
    fpr_max: float = 0.15,
) -> Dict[str, Any]:
    df = load_results_csv(csv_path)
    _require_cols(df, [region_col, true_col, pred_col])
    if base_filter_expr:
        df = df.query(base_filter_expr).copy()
    per_region: Dict[str, Any] = {}
    passed_all = True
    for r in regions:
        sub = df[df[region_col] == r].copy()
        y_true = _to_binary(sub[true_col]) if len(sub) else np.array([], dtype=int)
        y_pred = _to_binary(sub[pred_col]) if len(sub) else np.array([], dtype=int)
        met = (detection_metrics(y_true, y_pred) if len(sub)
               else {"fpr": np.nan, "recall": np.nan, "precision": np.nan, "f1": np.nan})
        ok = (not np.isnan(met["fpr"])) and (met["fpr"] <= fpr_max)
        per_region[r] = {"n": int(len(sub)), "fpr": float(met["fpr"]) if not np.isnan(met["fpr"]) else np.nan, "passed_region": bool(ok)}
        passed_all = passed_all and ok
    fprs = [v["fpr"] for v in per_region.values()]
    return _result("T-04", passed_all, {"fpr_max_observed": float(np.nanmax(fprs)) if fprs else np.nan},
        {"per_region": per_region, "regions": list(regions), "base_filter": base_filter_expr,
         "target": {"fpr_max": fpr_max}})


def test_T05_prediction(
    csv_path: str,
    *,
    true_power_col: str,
    pred_power_col: str,
    filter_expr: Optional[str] = None,
    energy_frac_min: float = 0.75,
    energy_frac_max: float = 1.25,
    norm_err_max: float = 0.45,
) -> Dict[str, Any]:
    df = load_results_csv(csv_path)
    _require_cols(df, [true_power_col, pred_power_col])
    if filter_expr:
        df = df.query(filter_expr).copy()
    yt = pd.to_numeric(df[true_power_col], errors="coerce").fillna(0).to_numpy(dtype=float)
    yp = pd.to_numeric(df[pred_power_col], errors="coerce").fillna(0).to_numpy(dtype=float)
    met = {
        "energy_frac": energy_fraction(yt, yp),
        "norm_err":    norm_l1_error(yt, yp),
    }
    passed = energy_frac_min <= met["energy_frac"] <= energy_frac_max and met["norm_err"] <= norm_err_max    
    return _result("T-05", passed, met,
    {"n": int(len(df)), "filter": filter_expr,
     "targets": {"energy_frac_min": energy_frac_min, "energy_frac_max": energy_frac_max, "norm_err_max": norm_err_max}})


def run_tests_on_file(csv_path: str, thresholds: Dict[str, Any]) -> Dict[str, Any]:
    """
    Exécute T-02, T-03, T-05 sur un fichier CSV.
    Dérive ac_on_true à partir de P_reel_climatisation > binarisation_threshold.
    """
    df = load_results_csv(csv_path)
    seuil = thresholds.get("binarisation_threshold", 0.0)
    df["ac_on_true"] = (df["P_reel_climatisation"] > seuil).astype(int)
    tmp = csv_path.replace(".csv", "_tmp_with_true.csv")
    df.to_csv(tmp, index=False)

    try:
        r_T02 = test_T02_detection(
            csv_path=tmp,
            true_col="ac_on_true",
            pred_col="o_climatisation",
            recall_min=thresholds["recall_min"],
            f1_min=thresholds["f1_min"],
        )
        r_T03 = test_T03_false_positives(
            csv_path=tmp,
            true_col="ac_on_true",
            pred_col="o_climatisation",
            fpr_max=thresholds["fpr_max"],
        )
        r_T05 = test_T05_prediction(
            csv_path=tmp,
            true_power_col="P_reel_climatisation",
            pred_power_col="P_estime_climatisation",
            energy_frac_min=thresholds["energy_frac_min"],
            energy_frac_max=thresholds["energy_frac_max"],  
            norm_err_max=thresholds["norm_err_max"],
        )
    finally:
        os.remove(tmp)

    return {"T02": r_T02, "T03": r_T03, "T05": r_T05}


def save_excel_report(
    all_results: list[Dict[str, Any]],
    out_path: str,
    thresholds: Dict[str, Any],
) -> None:

    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport de tests"

    # ── Styles ─────────────────────────────────────────────────────────────────
    header_font    = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill("solid", start_color="2F5496")
    subheader_fill = PatternFill("solid", start_color="D9E1F2")
    pass_fill      = PatternFill("solid", start_color="C6EFCE")
    fail_fill      = PatternFill("solid", start_color="FFC7CE")
    pass_font      = Font(name="Arial", bold=True, color="276221")
    fail_font      = Font(name="Arial", bold=True, color="9C0006")
    center         = Alignment(horizontal="center", vertical="center")
    thin           = Side(style="thin", color="BFBFBF")
    border         = Border(left=thin, right=thin, top=thin, bottom=thin)
    normal_font    = Font(name="Arial", size=10)
    desc_font      = Font(name="Arial", italic=True, size=9, color="595959")
    desc_align     = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Contenu ────────────────────────────────────────────────────────────────
    headers = [
        "Fichier CSV", "N",
        "T-02 Résultat", "Recall", "Precision", "F1",
        "T-03 Résultat", "FPR",
        "T-05 Résultat", "Energy Frac", "Norm Err",
        "Résultat global",
    ]

    seuils_row = [
        "— Seuils —", "—",
        "—", f"≥ {thresholds['recall_min']:.0%}", "—", f"≥ {thresholds['f1_min']:.0%}",
        "—", f"≤ {thresholds['fpr_max']:.0%}",
        "—", f"≥ {thresholds['energy_frac_min']:.0%} – {thresholds['energy_frac_max']:.0%}", f"≤ {thresholds['norm_err_max']:.0%}",
        "—",
    ]

    descriptions = {
        "T-02": (
            "Mesure la capacité de l'algorithme à détecter correctement les instants où la "
            "climatisation est allumée (ON), en évaluant le rappel (proportion de vrais ON "
            "détectés) et le score F1 (équilibre entre détection et précision)."
        ),
        "T-03": (
            "Mesure la proportion de pas de temps où la climatisation est réellement éteinte "
            "(OFF) mais que l'algorithme prédit à tort comme allumée (faux positifs)."
        ),
        "T-05": (
            "Mesure la qualité de l'estimation de la puissance consommée par la climatisation : "
            "la fraction d'énergie évalue si l'énergie totale estimée correspond à la réalité, "
            "et l'erreur normalisée quantifie l'écart cumulé entre la puissance estimée et la "
            "puissance réelle."
        ),
    }

    # ── Ligne 1 : Titre ────────────────────────────────────────────────────────
    ws.merge_cells("A1:L1")
    ws["A1"] = (
        f"Rapport de tests NILM — Climatisation  |  "
        f"Seuil ON : P_reel > {thresholds['binarisation_threshold']} kW"
    )
    ws["A1"].font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", start_color="1F3864")
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    # ── Ligne 2 : En-têtes ─────────────────────────────────────────────────────
    for col_idx, h in enumerate(headers, start=1):
        cell           = ws.cell(row=2, column=col_idx, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border

    # ── Ligne 3 : Seuils ───────────────────────────────────────────────────────
    for col_idx, s in enumerate(seuils_row, start=1):
        cell           = ws.cell(row=3, column=col_idx, value=s)
        cell.font      = Font(name="Arial", bold=True, size=9, color="2F5496")
        cell.fill      = subheader_fill
        cell.alignment = center
        cell.border    = border

    # # ── Ligne 4 : Descriptions ─────────────────────────────────────────────────
    # # Étape 1 : fusionner les cellules par groupe de test
    # ws.merge_cells("C4:F4")   # T-02
    # ws.merge_cells("G4:H4")   # T-03
    # ws.merge_cells("I4:K4")   # T-05

    # # Étape 2 : appliquer le style sur toutes les cellules de la ligne
    # for col_idx in range(1, 13):
    #     cell           = ws.cell(row=4, column=col_idx)
    #     cell.font      = desc_font
    #     cell.alignment = desc_align
    #     cell.border    = border

    # # Étape 3 : écrire le texte dans la cellule maîtresse de chaque groupe
    # ws.cell(row=4, column=3).value = descriptions["T-02"]
    # ws.cell(row=4, column=7).value = descriptions["T-03"]
    # ws.cell(row=4, column=9).value = descriptions["T-05"]

    # ws.row_dimensions[4].height = 60

    # ── Lignes de données ──────────────────────────────────────────────────────
    for entry in all_results:
        fname   = entry["filename"]
        T02     = entry["results"]["T02"]
        T03     = entry["results"]["T03"]
        T05     = entry["results"]["T05"]
        n       = T02["details"]["n"]
        overall = T02["passed"] and T03["passed"] and T05["passed"]

        row = [
            fname, n,
            "PASS" if T02["passed"] else "FAIL",
            round(T02["metrics"].get("recall",      float("nan")), 4),
            round(T02["metrics"].get("precision",   float("nan")), 4),
            round(T02["metrics"].get("f1",          float("nan")), 4),
            "PASS" if T03["passed"] else "FAIL",
            round(T03["metrics"].get("fpr",         float("nan")), 4),
            "PASS" if T05["passed"] else "FAIL",
            round(T05["metrics"].get("energy_frac", float("nan")), 4),
            round(T05["metrics"].get("norm_err",    float("nan")), 4),
            "PASS" if overall else "FAIL",
        ]
        ws.append(row)
        r = ws.max_row

        for col_idx in range(1, 13):
            cell           = ws.cell(row=r, column=col_idx)
            cell.font      = normal_font
            cell.alignment = center
            cell.border    = border

        for col_idx in [3, 7, 9, 12]:
            cell = ws.cell(row=r, column=col_idx)
            if cell.value == "PASS":
                cell.fill = pass_fill
                cell.font = pass_font
            else:
                cell.fill = fail_fill
                cell.font = fail_font

    # ── Largeurs de colonnes ───────────────────────────────────────────────────
    col_widths = [40, 8, 14, 10, 12, 10, 14, 10, 14, 13, 12, 16]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"

    # ── Feuille 2 : Descriptions des tests ────────────────────────────────────
    ws2 = wb.create_sheet(title="Description des tests")

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 100

    ws2.merge_cells("A1:B1")
    ws2["A1"] = "Description des tests"
    ws2["A1"].font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws2["A1"].fill      = PatternFill("solid", start_color="1F3864")
    ws2["A1"].alignment = center
    ws2.row_dimensions[1].height = 28

    for col_idx, h in enumerate(["Test", "Description"], start=1):
        cell           = ws2.cell(row=2, column=col_idx, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border

    desc_rows = [
        ("T-02", descriptions["T-02"]),
        ("T-03", descriptions["T-03"]),
        ("T-05", descriptions["T-05"]),
    ]
    for row_idx, (test_id, desc) in enumerate(desc_rows, start=3):
        cell_id              = ws2.cell(row=row_idx, column=1, value=test_id)
        cell_id.font         = Font(name="Arial", bold=True, size=10)
        cell_id.alignment    = center
        cell_id.border       = border

        cell_desc            = ws2.cell(row=row_idx, column=2, value=desc)
        cell_desc.font       = Font(name="Arial", size=10)
        cell_desc.alignment  = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell_desc.border     = border
        ws2.row_dimensions[row_idx].height = 60
    wb.save(out_path)


if __name__ == "__main__":

    # ── Configuration ──────────────────────────────────────────────────────────
    DOSSIER_CSV = "/Users/catherinehenri/Desktop/Genie Elec/Session H2026/ELE8080/Dev/Resultats"
    RAPPORT_OUT = "/Users/catherinehenri/Desktop/Genie Elec/Session H2026/ELE8080/Dev/Resultats/rapport_tests.xlsx"

    THRESHOLDS = {
        "recall_min":             0.70,
        "f1_min":                 0.70,
        "fpr_max":                0.15,
        "energy_frac_min":        0.75,
        "energy_frac_max":        1.25,
        "norm_err_max":           0.45,
        "binarisation_threshold": 0.15,   
    }
    # ──────────────────────────────────────────────────────────────────────────

    csv_files = glob.glob(os.path.join(DOSSIER_CSV, "*.csv"))

    if not csv_files:
        print(f"Aucun fichier CSV trouvé dans : {DOSSIER_CSV}")
    else:
        all_results = []
        for csv_path in sorted(csv_files):
            fname = os.path.basename(csv_path)
            print(f"  Traitement : {fname} ... ", end="")
            try:
                results = run_tests_on_file(csv_path, THRESHOLDS)
                all_results.append({"filename": fname, "results": results})
                overall = all(r["passed"] for r in results.values())
                print("PASS" if overall else "FAIL")
            except Exception as e:
                print(f"ERREUR : {e}")

        save_excel_report(all_results, RAPPORT_OUT, THRESHOLDS)
        print(f"\nRapport sauvegardé : {RAPPORT_OUT}")